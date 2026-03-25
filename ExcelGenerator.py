import os
import re
import time
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


start_time = time.perf_counter()

SOURCE_FOLDER = Path("../SOURCE FILES/")
EXCEL_OUTPUT_ROOT = Path("../EXCEL FILES/")
COMBINED_EXCEL_FILENAME = "AllProjects.xlsx"
COMBINED_SHEET_NAME = "ProtonView"
LANGUAGE_COLUMN_PATTERN = r"^[a-z]{2}-[A-Z]{2}$"
KEY_COLUMN_NAME = "Keys"
PROTONVIEW_COLUMN_NAME = "ProtonView"
PROTONVIEW_COLUMN_VALUE = "ProtonView_Resources_Controllers.IPlcResource"
EXCLUDED_LANGUAGE_COLUMNS = {"ta-GG"}


def load_csv_files() -> dict[str, pl.DataFrame]:
    csv_files = sorted(SOURCE_FOLDER.glob("*.csv"))
    dfs: dict[str, pl.DataFrame] = {}

    for path in csv_files:
        sheet_name = path.stem

        try:
            df = pl.read_csv(path)

            if df.height == 0 or df.columns is None or any(col is None for col in df.columns):
                print(f"Warning: file '{sheet_name}' has invalid or missing headers, skipping...")
                continue

            dfs[sheet_name] = df
            print(
                f"File '{sheet_name}' read successfully with "
                f"{df.height} rows and {df.width} columns."
            )
        except Exception as exc:
            print(f"Warning: error processing '{sheet_name}': {exc}")

    return dfs


def filter_dataframes(dfs: dict[str, pl.DataFrame]) -> dict[str, pl.DataFrame]:
    df_filtered_dict: dict[str, pl.DataFrame] = {}

    for sheet_name, df in dfs.items():
        if KEY_COLUMN_NAME not in df.columns:
            raise ValueError(
                f"File '{sheet_name}' must contain a '{KEY_COLUMN_NAME}' column."
            )

        key_column_index = df.columns.index(KEY_COLUMN_NAME)
        translation_columns = [
            column_name
            for column_name in df.columns[key_column_index + 1:]
            if column_name not in EXCLUDED_LANGUAGE_COLUMNS
        ]

        if not translation_columns:
            raise ValueError(
                f"File '{sheet_name}' must contain at least one language column after '{KEY_COLUMN_NAME}'."
            )

        df = df.with_columns(
            [pl.col(column).fill_null("").cast(pl.Utf8) for column in translation_columns]
        )

        base_language_column = translation_columns[0]

        df_filtered = df.filter(
            ~pl.col(base_language_column).str.contains(r"(?i).*Undefined")
        )

        df_filtered2 = df_filtered.filter(
            ~pl.col(base_language_column).str.contains(r"(?i)Unkown message")
        )

        df_filtered_subset = df_filtered2.select([KEY_COLUMN_NAME] + translation_columns)

        df_cleaned = df_filtered_subset.filter(
            pl.col(base_language_column) != ""
        )

        df_filtered_dict[sheet_name] = df_cleaned

        print("=" * 50)
        print(f"Sheet: {sheet_name}")
        print(f"Original length: {len(df)}")
        print(f"Filtered length: {len(df_cleaned)}")
        print(f"Data has been reduced by a {100 - len(df_cleaned) * 100 / len(df):.2f}%")
        print("=" * 50)

    return df_filtered_dict


def validate_language_columns(df_filtered_dict: dict[str, pl.DataFrame]) -> None:
    for sheet_name, df in df_filtered_dict.items():
        for column_name in df.columns[1:]:
            if not re.match(LANGUAGE_COLUMN_PATTERN, column_name):
                raise ValueError(
                    f"\nColumn name '{column_name}' of {sheet_name} does not match "
                    "format [a-z][a-z]-[A-Z][A-Z]"
                )


def prefix_keys_with_project(df_filtered_dict: dict[str, pl.DataFrame]) -> dict[str, pl.DataFrame]:
    prefixed_dict: dict[str, pl.DataFrame] = {}

    for sheet_name, df in df_filtered_dict.items():
        project_name = get_project_name(sheet_name)
        key_column = df.columns[0]
        key_prefix = f"{project_name}_"

        prefixed_dict[sheet_name] = df.with_columns(
            pl.when(pl.col(key_column).cast(pl.Utf8).str.starts_with(key_prefix))
            .then(pl.col(key_column).cast(pl.Utf8))
            .otherwise(pl.lit(key_prefix) + pl.col(key_column).cast(pl.Utf8))
            .alias(key_column)
        )

    return prefixed_dict


def get_all_language_columns(df_filtered_dict: dict[str, pl.DataFrame]) -> list[str]:
    return sorted(
        {
            column_name
            for df in df_filtered_dict.values()
            for column_name in df.columns[1:]
            if column_name not in EXCLUDED_LANGUAGE_COLUMNS
        }
    )


def normalize_language_columns(df_filtered_dict: dict[str, pl.DataFrame]) -> dict[str, pl.DataFrame]:
    normalized_dict: dict[str, pl.DataFrame] = {}
    all_language_columns = get_all_language_columns(df_filtered_dict)

    for sheet_name, df in df_filtered_dict.items():
        key_column = df.columns[0]
        existing_language_columns = df.columns[1:]
        fallback_language = "en-US" if "en-US" in existing_language_columns else existing_language_columns[0]

        missing_column_exprs = [
            pl.col(fallback_language).cast(pl.Utf8).alias(column_name)
            for column_name in all_language_columns
            if column_name not in existing_language_columns
        ]

        if missing_column_exprs:
            df = df.with_columns(missing_column_exprs)

        normalized_dict[sheet_name] = df.select([key_column] + all_language_columns)

    return normalized_dict


def prepend_protonview_column(df_filtered_dict: dict[str, pl.DataFrame]) -> dict[str, pl.DataFrame]:
    protonview_dict: dict[str, pl.DataFrame] = {}

    for sheet_name, df in df_filtered_dict.items():
        protonview_dict[sheet_name] = df.with_columns(
            pl.lit(PROTONVIEW_COLUMN_VALUE).alias(PROTONVIEW_COLUMN_NAME)
        ).select([PROTONVIEW_COLUMN_NAME] + df.columns)

    return protonview_dict


def get_project_name(sheet_name: str) -> str:
    return sheet_name.split("_", 1)[0]


def autosize_columns(worksheet) -> None:
    for column_index, column_cells in enumerate(worksheet.iter_cols(), start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max(max_length + 2, 12), 80)
        worksheet.column_dimensions[get_column_letter(column_index)].width = adjusted_width


def write_rows_to_sheet(
    worksheet,
    start_row: int,
    df: pl.DataFrame,
) -> int:
    header_row = start_row
    for column_index, column_name in enumerate(df.columns, start=1):
        worksheet.cell(row=header_row, column=column_index, value=column_name)

    if df.height == 0:
        return header_row + 3

    for row_index, row in enumerate(df.iter_rows(), start=header_row + 1):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    return header_row + df.height + 3


def clear_legacy_project_workbooks(project_names: set[str]) -> None:
    for project_name in project_names:
        legacy_path = EXCEL_OUTPUT_ROOT / f"{project_name}.xlsx"
        if legacy_path.exists():
            legacy_path.unlink()
            print(f"Removed legacy Excel file: {legacy_path.as_posix()}")


def generate_excel_files(df_filtered_dict: dict[str, pl.DataFrame]) -> None:
    EXCEL_OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = COMBINED_SHEET_NAME
    current_row = 1

    ordered_items = sorted(
        df_filtered_dict.items(),
        key=lambda item: (get_project_name(item[0]), item[0]),
    )

    if ordered_items:
        combined_df = pl.concat(
            [df for _, df in ordered_items],
            how="vertical",
        )
        current_row = write_rows_to_sheet(
            worksheet,
            current_row,
            combined_df,
        )

    autosize_columns(worksheet)

    clear_legacy_project_workbooks({get_project_name(sheet_name) for sheet_name in df_filtered_dict})

    output_path = EXCEL_OUTPUT_ROOT / COMBINED_EXCEL_FILENAME
    temp_output_path = output_path.with_suffix(".tmp.xlsx")

    try:
        workbook.save(temp_output_path)
        os.replace(temp_output_path, output_path)
    except PermissionError as exc:
        if temp_output_path.exists():
            temp_output_path.unlink(missing_ok=True)
        raise PermissionError(
            f"Cannot overwrite '{output_path}'. Close the workbook if it is open and run the generator again."
        ) from exc
    finally:
        workbook.close()

    print(f"Excel file created: {output_path.as_posix()}")


def main() -> None:
    dfs = load_csv_files()
    print(f"Current time {time.perf_counter() - start_time:.2f} seconds")

    df_filtered_dict = filter_dataframes(dfs)
    validate_language_columns(df_filtered_dict)
    df_filtered_dict = prefix_keys_with_project(df_filtered_dict)
    df_filtered_dict = normalize_language_columns(df_filtered_dict)
    df_filtered_dict = prepend_protonview_column(df_filtered_dict)
    generate_excel_files(df_filtered_dict)

    elapsed = time.perf_counter() - start_time
    print(f"\nTotal execution time: {elapsed:.2f} seconds")


if __name__ == "__main__":
    main()
